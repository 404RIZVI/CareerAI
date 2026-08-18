import json
import os
from datetime import datetime

from flask import Flask, render_template, request, redirect, url_for, session, flash

from openpyxl import Workbook, load_workbook

from database import (
    init_db,
    get_careers,
    get_career_by_name,
    get_questions,
    save_result
)

from recommendation import (
    recommend,
    skill_gap,
    compare_careers
)

from ai_service import generate_explanation


app = Flask(__name__)

app.config["SECRET_KEY"] = "hackathon-career-guidance-dev-key"
app.config["MAX_CONTENT_LENGTH"] = 1 * 1024 * 1024

# Result ko temporarily server memory me rakhenge
LATEST_RESULT = result = None


# =========================================================
# DATABASE
# =========================================================

init_db()


# =========================================================
# EXCEL FILE
# =========================================================

EXCEL_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "students_data.xlsx"
)


# =========================================================
# STREAM SUBJECTS
# =========================================================

STREAM_SUBJECTS = {
    "Science": [
        "Mathematics",
        "Physics",
        "Chemistry",
        "Computer Science"
    ],

    "Commerce": [
        "Mathematics",
        "Accountancy",
        "Economics",
        "Business Studies",
        "Computer Science"
    ],

    "Arts": [
        "English",
        "History",
        "Political Science",
        "Psychology",
        "Economics"
    ],
}


# =========================================================
# INTERESTS
# =========================================================

INTERESTS = [
    "Artificial Intelligence",
    "Programming",
    "Data",
    "Finance",
    "Business",
    "Cyber Security",
    "Design",
    "Robotics",
    "Cloud Computing",
    "Research",
    "Marketing",
    "Management"
]


# =========================================================
# SKILLS
# =========================================================

SKILLS = [
    "Python",
    "C/C++",
    "JavaScript",
    "HTML/CSS",
    "SQL",
    "Excel",
    "Data Analysis",
    "Mathematics",
    "Communication",
    "Design",
    "Problem Solving",
    "Leadership"
]


# =========================================================
# PREFERENCES
# =========================================================

PREFERENCE_FIELDS = [
    ("coding", "I enjoy coding and building software"),
    ("math", "I enjoy mathematics and logical reasoning"),
    ("data", "I like working with data and finding patterns"),
    ("creative", "I prefer creative/design-oriented work"),
    ("technical", "I enjoy solving technical problems"),
    ("business", "I enjoy business, finance or strategy"),
    ("research", "I like research and experimentation"),
    ("physical", "I like building physical/robotic systems"),
]


# =========================================================
# FORM VALIDATION
# =========================================================

def validate_form(form):

    name = form.get("name", "").strip()
    age_raw = form.get("age", "").strip()
    education = form.get("education", "").strip()
    stream = form.get("stream", "").strip()

    if not name or not age_raw or not education or not stream:
        return None, "Please complete all basic profile fields."

    try:
        age = int(age_raw)
    except ValueError:
        return None, "Age must be a whole number."

    if age < 13 or age > 100:
        return None, "Please enter a valid age."

    # -----------------------------
    # SUBJECT MARKS
    # -----------------------------

    subjects = {}

    for subject in STREAM_SUBJECTS.get(stream, []):

        raw = form.get(
            "subject_" + subject,
            ""
        ).strip()

        if raw == "":
            return None, f"Please enter a mark for {subject}."

        try:
            mark = float(raw)
        except ValueError:
            return None, f"{subject} marks must be numeric."

        if mark < 0 or mark > 100:
            return None, f"{subject} marks must be between 0 and 100."

        subjects[subject] = mark

    # -----------------------------
    # INTERESTS
    # -----------------------------

    interests = form.getlist("interests")

    if not interests:
        return None, "Select at least one interest."

    # -----------------------------
    # SKILLS
    # -----------------------------

    skills = form.getlist("skills")

    if not skills:
        return None, "Select at least one skill."

    # -----------------------------
    # PREFERENCES
    # -----------------------------

    preferences = {}

    for key, _ in PREFERENCE_FIELDS:

        raw = form.get(key, "")

        if raw not in {"1", "2", "3", "4", "5"}:
            return None, "Please answer all career preference questions."

        preferences[key] = int(raw)

    # -----------------------------
    # APTITUDE
    # -----------------------------

    aptitude = []

    for i in range(1, 11):

        raw = form.get(
            f"aptitude_{i}",
            ""
        )

        if raw not in {"0", "1", "2", "3", "4"}:
            return None, "Please answer all assessment questions."

        aptitude.append(int(raw))

    return {
        "name": name,
        "age": age,
        "education": education,
        "stream": stream,
        "subjects": subjects,
        "interests": interests,
        "skills": skills,
        "preferences": preferences,
        "aptitude": aptitude,
    }, None


# =========================================================
# SAVE STUDENT DATA TO EXCEL
# =========================================================

def save_student_to_excel(answers, result):

    try:

        # Existing Excel file open karo
        if os.path.exists(EXCEL_FILE):

            workbook = load_workbook(EXCEL_FILE)

            if "Student Results" in workbook.sheetnames:
                sheet = workbook["Student Results"]
            else:
                sheet = workbook.create_sheet("Student Results")

            # Ensure the header exists even if the workbook was created earlier.
            if sheet.max_row == 1 and all(
                cell.value is None for cell in sheet[1]
            ):
                sheet.append([
                    "Date & Time",
                    "Name",
                    "Age",
                    "Education",
                    "Stream",
                    "Subjects & Marks",
                    "Interests",
                    "Skills",
                    "Best Career",
                    "Match %",
                    "Top 3 Careers",
                    "Missing Skills"
                ])

        # Agar file nahi hai to new Excel banao
        else:

            workbook = Workbook()

            sheet = workbook.active

            sheet.title = "Student Results"

            sheet.append([
                "Date & Time",
                "Name",
                "Age",
                "Education",
                "Stream",
                "Subjects & Marks",
                "Interests",
                "Skills",
                "Best Career",
                "Match %",
                "Top 3 Careers",
                "Missing Skills"
            ])

        # Subjects ko text me convert karo
        subjects_text = ", ".join(
            f"{subject}: {mark}"
            for subject, mark in answers["subjects"].items()
        )

        # Top 3 careers
        top3_text = " | ".join(
            f"{item['career']['name']} - {item['score']}%"
            for item in result["top3"]
        )

        # Excel me new row
        sheet.append([
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),

            answers["name"],

            answers["age"],

            answers["education"],

            answers["stream"],

            subjects_text,

            ", ".join(
                answers["interests"]
            ),

            ", ".join(
                answers["skills"]
            ),

            result["top"]["career"]["name"],

            result["top"]["score"],

            top3_text,

            ", ".join(
                result["skill_gap"]["missing"]
            )
        ])

        workbook.save(EXCEL_FILE)

        print()
        print("================================")
        print("Excel data saved successfully!")
        print("File:", EXCEL_FILE)
        print("Student:", answers["name"])
        print("================================")
        print()

    except Exception as error:

        print("Excel save error:", error)


# =========================================================
# HOME
# =========================================================

@app.route("/")
def index():

    return render_template(
        "index.html"
    )


# =========================================================
# ASSESSMENT
# =========================================================

@app.route("/assessment")
def assessment():

    return render_template(
        "assessment.html",
        subjects=STREAM_SUBJECTS,
        interests=INTERESTS,
        skills=SKILLS,
        preference_fields=PREFERENCE_FIELDS,
        questions=get_questions()
    )


# =========================================================
# ANALYZE
# =========================================================

@app.post("/analyze")
def analyze():

    global LATEST_RESULT

    # Form validate karo
    answers, error = validate_form(
        request.form
    )

    if error:

        return (
            f"Assessment Error: {error}",
            400
        )

    # Careers lao
    careers = get_careers()

    # Recommendation engine
    top3, all_scores = recommend(
        careers,
        answers
    )

    if not top3:

        return (
            "No career recommendations found.",
            500
        )

    # Best career
    top = top3[0]

    # Skill gap
    gap = skill_gap(
        top["career"],
        answers["skills"]
    )

    # AI ke liye profile
    profile_for_ai = {
        "education": answers["education"],
        "stream": answers["stream"],
        "interests": answers["interests"],
        "skills": answers["skills"],
        "subjects": answers["subjects"],
        "skill_gap": gap,
    }

    # AI explanation
    explanation, ai_used = generate_explanation(
        top,
        profile_for_ai
    )

    # Final result
    result = {
        "answers": answers,
        "top3": top3,
        "all_scores": all_scores,
        "top": top,
        "skill_gap": gap,
        "explanation": explanation,
        "ai_used": ai_used,
    }

    # IMPORTANT:
    # Session ki jagah server memory me result save
    LATEST_RESULT = result

    # Database save
    save_result(
        answers["name"],
        result
    )

    # Excel save
    save_student_to_excel(
        answers,
        result
    )

    # Terminal confirmation
    print()
    print("================================")
    print("CAREER ANALYSIS COMPLETE")
    print("Student:", answers["name"])
    print(
        "Best Career:",
        top["career"]["name"]
    )
    print(
        "Match:",
        top["score"],
        "%"
    )
    print("================================")
    print()

    # Result page par bhejo
    return redirect(
        url_for("result")
    )


# =========================================================
# RESULT
# =========================================================

@app.route("/result")
def result():

    global LATEST_RESULT

    if LATEST_RESULT is None:

        return redirect(
            url_for("assessment")
        )

    print(
        "Showing result for:",
        LATEST_RESULT["answers"]["name"]
    )

    return render_template(
        "result.html",
        result=LATEST_RESULT
    )


# =========================================================
# COMPARE
# =========================================================

@app.route("/compare")
def compare():

    global LATEST_RESULT

    if LATEST_RESULT is None:

        return redirect(
            url_for("assessment")
        )

    careers = get_careers()

    left_name = request.args.get(
        "left",
        LATEST_RESULT["top3"][0]["career"]["name"]
    )

    right_name = request.args.get(
        "right",
        LATEST_RESULT["top3"][1]["career"]["name"]
    )

    left = (
        get_career_by_name(left_name)
        or careers[0]
    )

    right = (
        get_career_by_name(right_name)
        or careers[1]
    )

    comparison = compare_careers(
        left,
        right,
        LATEST_RESULT["answers"]
    )

    return render_template(
        "compare.html",
        careers=careers,
        comparison=comparison
    )


# =========================================================
# WHAT IF
# =========================================================

@app.route("/what-if/<path:career_name>")
def what_if(career_name):
    global LATEST_RESULT

    if LATEST_RESULT is None:

        return redirect(
            url_for("assessment")
        )

    career = get_career_by_name(
        career_name
    )

    if not career:

        return render_template(
            "error.html",
            message="Career not found."
        ), 404

    from recommendation import (
        score_career,
        reasons_for
    )

    score, parts = score_career(
        career,
        LATEST_RESULT["answers"]
    )

    gap = skill_gap(
        career,
        LATEST_RESULT["answers"]["skills"]
    )

    result_data = {
        **LATEST_RESULT,

        "top": {
            "career": career,
            "score": score,
            "parts": parts,
            "reasons": reasons_for(
                career,
                LATEST_RESULT["answers"],
                parts
            )
        },

        "skill_gap": gap,

        "explanation": (
            f"If you choose {career['name']}, "
            f"your current estimated match is {score}%. "
            f"Focus next on "
            f"{', '.join(gap['missing'][:4]) or 'deeper projects and specialization'}."
        ),

        "ai_used": False,

        "what_if": True,
    }

    return render_template(
        "result.html",
        result=result_data
    )


# =========================================================
# ERROR HANDLERS
# =========================================================

@app.errorhandler(413)
def too_large(error):

    return render_template(
        "error.html",
        message=(
            "Submission is too large. "
            "Please keep the assessment concise."
        )
    ), 413


@app.errorhandler(500)
def server_error(error):

    return render_template(
        "error.html",
        message=(
            "Something went wrong on the server. "
            "Please try the assessment again."
        )
    ), 500


# =========================================================
# START FLASK
# =========================================================

if __name__ == "__main__":

    app.run(
        debug=True
    )
